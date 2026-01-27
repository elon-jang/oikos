"""헌금 전표 Excel 처리 스크립트.

Functions:
    create_template(date_str) — 샘플 xlsx 복사 → 날짜 세팅 → YYYYMMDD.xlsx 생성
    write_data(date_str, data_json) — JSON 데이터를 Excel에 입력
    verify(date_str) — 입력된 데이터 읽어서 카테고리별 합계 출력
"""

import json
import os
import shutil
import sys
from datetime import datetime

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
TEMPLATE_FILE = os.path.join(BASE_DIR, "templates", "upload_sample.xlsx")

# Import config
sys.path.insert(0, SCRIPT_DIR)
from offering_config import CATEGORIES, CATEGORY_ORDER


def _parse_date(date_str: str) -> datetime:
    """YYYYMMDD 문자열을 datetime으로 변환."""
    return datetime.strptime(date_str, "%Y%m%d")


def _backup_file(filepath: str) -> str:
    """기존 파일을 _backup.xlsx로 백업. 백업 경로 반환."""
    backup_path = filepath.replace(".xlsx", "_backup.xlsx")
    shutil.copy2(filepath, backup_path)
    print(f"기존 파일 백업: {backup_path}")
    return backup_path


def _output_path(date_str: str) -> str:
    """날짜에 해당하는 Excel 파일 경로."""
    year = date_str[:4]
    mmdd = date_str[4:]
    data_dir = os.path.join(BASE_DIR, "data", year, mmdd)
    os.makedirs(data_dir, exist_ok=True)
    return os.path.join(data_dir, f"{date_str}.xlsx")


def create_template(date_str: str) -> str:
    """샘플 xlsx를 복사하고 날짜를 세팅하여 새 파일 생성.

    Returns:
        생성된 파일 경로
    """
    output_path = _output_path(date_str)
    target_date = _parse_date(date_str)

    # 이미 파일이 있으면 백업
    if os.path.exists(output_path):
        _backup_file(output_path)

    # 템플릿 복사
    shutil.copy2(TEMPLATE_FILE, output_path)

    # 날짜 세팅
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    first_row = min(c["start"] for c in CATEGORIES.values())
    last_row = max(c["end"] for c in CATEGORIES.values())
    for row_num in range(first_row, last_row + 1):
        cell_a = ws.cell(row=row_num, column=1)  # Column A: 일자
        cell_a.value = target_date

    wb.save(output_path)
    print(f"템플릿 생성 완료: {output_path}")
    return output_path


def _validate_data(data: dict) -> list[str]:
    """데이터 구조 검증. 오류 메시지 목록 반환."""
    errors = []
    if not isinstance(data, dict):
        return [f"데이터가 dict 형식이어야 합니다 (현재: {type(data).__name__})"]
    for category, entries in data.items():
        if not isinstance(entries, list):
            errors.append(f"{category}: 항목이 list여야 합니다 (현재: {type(entries).__name__})")
            continue
        for i, entry in enumerate(entries):
            if not isinstance(entry, dict):
                errors.append(f"{category}[{i}]: 항목이 dict여야 합니다")
                continue
            if "name" not in entry or not str(entry["name"]).strip():
                errors.append(f"{category}[{i}]: 이름이 비어있습니다")
            if "amount" not in entry:
                errors.append(f"{category}[{i}]: 금액(amount)이 없습니다")
            elif not isinstance(entry["amount"], (int, float)) or entry["amount"] < 0:
                errors.append(f"{category}[{i}]: 금액은 0 이상의 숫자여야 합니다 (현재: {entry['amount']})")
    return errors


def write_data(date_str: str, data_json: str, dry_run: bool = False) -> str:
    """JSON 데이터를 Excel에 입력.

    Args:
        date_str: YYYYMMDD 형식 날짜
        data_json: 카테고리별 데이터 JSON 문자열
            {
                "십일조": [{"name": "최정호", "amount": 578000}, ...],
                "장년부주일": [{"name": "장년부", "amount": 242000}],
                ...
            }
        dry_run: True면 검증+미리보기만 수행, 파일 쓰기 안함

    Returns:
        처리 결과 메시지
    """
    data = json.loads(data_json) if isinstance(data_json, str) else data_json

    # 입력 데이터 검증
    validation_errors = _validate_data(data)
    if validation_errors:
        msg = "❌ 데이터 검증 실패:\n" + "\n".join(validation_errors)
        print(msg)
        return msg

    warnings = []
    total_entries = 0
    grand_total = 0
    category_summaries = []

    # 중복 감지 (동일 카테고리 내 이름+금액 중복)
    for category, entries in data.items():
        seen = {}
        for entry in entries:
            key = (entry.get("name", ""), entry.get("amount", 0))
            if key in seen:
                warnings.append(
                    f"⚠ 중복 항목 감지 ({category}): {key[0]} {key[1]:,}원"
                )
            else:
                seen[key] = True

    for category, entries in data.items():
        if category not in CATEGORIES:
            warnings.append(f"⚠ 알 수 없는 카테고리: {category}")
            continue

        config = CATEGORIES[category]
        slots = config["slots"]

        if len(entries) > slots:
            warnings.append(
                f"⚠ {category}: {len(entries)}건 > {slots}슬롯 (초과 {len(entries) - slots}건 누락됨)"
            )

        effective = entries[:slots]
        cat_total = sum(e.get("amount", 0) for e in effective)
        total_entries += len(effective)
        grand_total += cat_total
        category_summaries.append(f"  {category}: {len(effective)}건, {cat_total:,}원")

    if dry_run:
        lines = [f"[dry-run] 검증 통과 — {total_entries}건, 합계 {grand_total:,}원"]
        lines.extend(category_summaries)
        if warnings:
            lines.extend(warnings)
        result = "\n".join(lines)
        print(result)
        return result

    # 실제 파일 쓰기
    output_path = _output_path(date_str)
    if not os.path.exists(output_path):
        create_template(date_str)
    else:
        _backup_file(output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    for category, entries in data.items():
        if category not in CATEGORIES:
            continue
        config = CATEGORIES[category]
        start_row = config["start"]
        slots = config["slots"]

        for i, entry in enumerate(entries[:slots]):
            row_num = start_row + i
            ws.cell(row=row_num, column=5).value = entry.get("name", "")
            ws.cell(row=row_num, column=7).value = entry.get("amount", 0)

    wb.save(output_path)

    result = f"입력 완료: {total_entries}건 → {output_path}"
    if warnings:
        result += "\n" + "\n".join(warnings)
    print(result)
    return result


def verify(date_str: str) -> str:
    """입력된 데이터를 읽어서 카테고리별 합계 출력.

    Returns:
        검증 결과 JSON 문자열
    """
    output_path = _output_path(date_str)
    if not os.path.exists(output_path):
        return json.dumps({"error": f"파일 없음: {output_path}"}, ensure_ascii=False)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    result = {}
    grand_total = 0

    for category in CATEGORY_ORDER:
        config = CATEGORIES[category]
        start_row = config["start"]
        end_row = config["end"]

        entries = []
        cat_total = 0

        for row_num in range(start_row, end_row + 1):
            name = ws.cell(row=row_num, column=5).value    # Column E
            amount = ws.cell(row=row_num, column=7).value  # Column G

            if name and amount:
                entries.append({"name": name, "amount": amount})
                cat_total += amount

        if entries:
            result[category] = {
                "entries": entries,
                "count": len(entries),
                "total": cat_total,
            }
            grand_total += cat_total

    result["_summary"] = {
        "grand_total": grand_total,
        "categories_with_data": len([k for k in result if not k.startswith("_")]),
    }

    output = json.dumps(result, ensure_ascii=False, indent=2)
    print(output)
    return output


def summary(month_str: str) -> str:
    """월간 요약 — 해당 월의 모든 주차 Excel을 집계.

    Args:
        month_str: YYYYMM 형식 (예: 202601)

    Returns:
        월간 요약 JSON 문자열
    """
    year = month_str[:4]
    month = month_str[4:6]
    year_dir = os.path.join(BASE_DIR, "data", year)

    if not os.path.isdir(year_dir):
        return json.dumps({"error": f"데이터 없음: {year_dir}"}, ensure_ascii=False)

    # MMDD 폴더 중 해당 월 필터링
    weekly_files = []
    for mmdd in sorted(os.listdir(year_dir)):
        mmdd_path = os.path.join(year_dir, mmdd)
        if not os.path.isdir(mmdd_path) or len(mmdd) != 4 or mmdd[:2] != month:
            continue
        xlsx = os.path.join(mmdd_path, f"{year}{mmdd}.xlsx")
        if os.path.exists(xlsx):
            weekly_files.append((mmdd, xlsx))

    if not weekly_files:
        return json.dumps(
            {"error": f"{year}년 {month}월 데이터 없음"}, ensure_ascii=False
        )

    # 카테고리별 월 합산
    monthly = {}
    weekly_totals = []

    for mmdd, xlsx_path in weekly_files:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        week_total = 0

        for category in CATEGORY_ORDER:
            config = CATEGORIES[category]
            for row_num in range(config["start"], config["end"] + 1):
                name = ws.cell(row=row_num, column=5).value
                amount = ws.cell(row=row_num, column=7).value
                if name and amount:
                    if category not in monthly:
                        monthly[category] = {"count": 0, "total": 0}
                    monthly[category]["count"] += 1
                    monthly[category]["total"] += amount
                    week_total += amount

        wb.close()
        weekly_totals.append({"date": mmdd, "total": week_total})

    grand_total = sum(w["total"] for w in weekly_totals)

    result = {
        "month": f"{year}-{month}",
        "weeks": len(weekly_files),
        "weekly_totals": weekly_totals,
        "categories": monthly,
        "grand_total": grand_total,
    }

    output = json.dumps(result, ensure_ascii=False, indent=2)
    print(output)
    return output


def rollback(date_str: str) -> str:
    """백업 파일로 Excel 복원.

    Args:
        date_str: YYYYMMDD 형식 날짜

    Returns:
        복원 결과 메시지
    """
    output_path = _output_path(date_str)
    backup_path = output_path.replace(".xlsx", "_backup.xlsx")

    if not os.path.exists(backup_path):
        msg = f"백업 파일 없음: {backup_path}"
        print(msg)
        return msg

    shutil.copy2(backup_path, output_path)
    os.remove(backup_path)
    msg = f"복원 완료: {backup_path} → {output_path}"
    print(msg)
    return msg


def main():
    """CLI 인터페이스.

    Usage:
        python process_offering.py create 20260125
        python process_offering.py write 20260125 < data.json
        python process_offering.py write --dry-run 20260125 < data.json
        python process_offering.py verify 20260125
        python process_offering.py rollback 20260125
        python process_offering.py summary 202601
    """
    if len(sys.argv) < 3:
        print("Usage: python process_offering.py <command> <date> [data_json]")
        print("Commands: create, write, verify, rollback, summary")
        sys.exit(1)

    command = sys.argv[1]

    if command == "summary":
        month_str = sys.argv[2]
        if len(month_str) == 2:
            month_str = f"{datetime.now().year}{month_str}"
        summary(month_str)
        return

    # write --dry-run 처리
    dry_run = False
    args = sys.argv[2:]
    if command == "write" and args and args[0] == "--dry-run":
        dry_run = True
        args = args[1:]

    if not args:
        print("Usage: python process_offering.py <command> <date> [data_json]")
        sys.exit(1)

    date_str = args[0]
    if len(date_str) == 4:
        date_str = f"{datetime.now().year}{date_str}"

    if command == "create":
        create_template(date_str)
    elif command == "write":
        if len(args) < 2:
            data_json = sys.stdin.read()
        else:
            data_json = args[1]
        write_data(date_str, data_json, dry_run=dry_run)
    elif command == "verify":
        verify(date_str)
    elif command == "rollback":
        rollback(date_str)
    else:
        print(f"Unknown command: {command}")
        sys.exit(1)


if __name__ == "__main__":
    main()
